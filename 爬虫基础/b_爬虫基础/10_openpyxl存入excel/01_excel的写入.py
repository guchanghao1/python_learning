# 导入 写入excel的工具
from openpyxl.workbook import Workbook

# 用列表 表示 一行
# li = ["姓名","年龄","身高"]
# name_list = ["江哥",18,"180"]  # 字符串默认是左对齐   数字默认是右对齐

data = [["姓名","年龄","身高"],["江哥",18,"180"],["立青",18,"170"]]

# 1、创建一个excel文件
wb = Workbook()

# 2、指定活动的工作表
sheet = wb.active
sheet.title = "表1"

# 3、往sheet表里面写数据
# sheet.append(li)
# sheet.append(name_list)
for i in data:
    print(i)
    sheet.append(i)



# 创建第二个工作表sheet2
xx = wb.create_sheet("这我第二个工作表")
xx.append(["Frank",18,"190"])


xxx = wb.create_sheet()
xxx.title = "第三个表"

# 4、保存
wb.save("test.xlsx")

