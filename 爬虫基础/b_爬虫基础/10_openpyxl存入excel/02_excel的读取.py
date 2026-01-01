from openpyxl import load_workbook  # load_workbook 方法

read_data = load_workbook("test.xlsx")
print(read_data)

# 1、指定读取哪张表
sheet = read_data.active

# 读取表1中的行
for 行 in sheet:
    for 单元格 in 行:
        print(单元格.value,end='   ')
    print()

# print(read_data["表1"])
# print(read_data["这我第二个工作表"])
# print(read_data["第三个表"])
#
# for hang in read_data["这我第二个工作表"]:
#     for danyuange in hang:
#         print(danyuange.value,end=" ")
#     print()